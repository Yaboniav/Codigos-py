import os
import pandas as pd
from pyxlsb import open_workbook as open_xlsb
from openpyxl import load_workbook

# Ruta de la carpeta con archivos .xlsb y .xlsx
folder_path = 'D://OneDrive - Grupo EPM//3. CENS//0. PIE - PIR - PF//3. Calidad del servicio//1. DB//2023'

# Lista para almacenar los datos
data = []

# Función para extraer encabezados de un archivo xlsb
def read_xlsb_headers(filename):
    with open_xlsb(filename) as wb:
        if "Exportar Hoja de Trabajo" in wb.sheets:
            with wb.get_sheet("Exportar Hoja de Trabajo") as sheet:
                first_row = next(sheet.rows())
                return [item.v for item in first_row if item.v is not None]
        else:
            return ["Hoja no encontrada"]

# Función para extraer encabezados de un archivo xlsx
def read_xlsx_headers(filename):
    wb = load_workbook(filename, read_only=True)
    if "Exportar Hoja de Trabajo" in wb.sheetnames:
        sheet = wb["Exportar Hoja de Trabajo"]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        return [cell for cell in first_row if cell is not None]
    else:
        return ["Hoja no encontrada"]

# Iterar sobre cada archivo en la carpeta
for filename in os.listdir(folder_path):
    if filename.endswith('.xlsb') or filename.endswith('.xlsx'):
        # Construir la ruta completa del archivo
        file_path = os.path.join(folder_path, filename)
        if filename.endswith('.xlsb'):
            headers = read_xlsb_headers(file_path)
        else:
            headers = read_xlsx_headers(file_path)
        headers.append(filename)  # Añadir el nombre del archivo al final de la lista
        data.append(headers)

# Crear un DataFrame con los datos, sin especificar las columnas por adelantado
df_titles = pd.DataFrame(data)

# Asignar nombres de columna al DataFrame: la última columna será 'Nombre del Libro'
df_titles.columns = [f'Columna {i}' for i in range(df_titles.shape[1] - 1)] + ['Nombre del Libro']

# Mostrar el DataFrame con los encabezados de cada archivo
print(df_titles)
